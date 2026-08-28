#!/usr/bin/env python3
"""
Track 1 - Families A & E: read the filled GT internal intake templates into summary JSON.

Reads the 'Data' sheet of an intake workbook (GT_Licensing_Startups_Intake.xlsx or
GT_Industry_Research_Intake.xlsx), pulls each metric's values by fiscal year, and writes
  output/internal_licensing.json   (Family A)
  output/internal_industry.json    (Family E)

Empty templates are fine — they produce a summary with no values, and the site cards stay
"In development" until real figures are entered. Once entered, re-run this then inject_downstream.py.

USAGE
  python3 read_internal.py ../templates/GT_Licensing_Startups_Intake.xlsx ../templates/GT_Industry_Research_Intake.xlsx
  python3 read_internal.py path/to/filled_template.xlsx
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

LICENSING_KEYS = {"invention_disclosures", "licenses_options_executed", "licensing_income_usd",
                  "startups_formed", "startups_active"}
INDUSTRY_KEYS = {"industry_sponsored_research_usd", "active_industry_agreements",
                 "corporate_affiliate_revenue_usd", "total_research_expenditures_usd"}
FAMILY_META = {
    "licensing": {"family": "licensing", "track": "Track 1 - Family A (licensing & startups)",
                  "source": "GT internal intake (Office of Technology Licensing / GTRC)"},
    "industry": {"family": "industry", "track": "Track 1 - Family E (industry research engagement)",
                 "source": "GT internal intake (GT Research / sponsored-programs accounting)"},
}


def find_header(ws):
    for r in range(1, 6):
        if str(ws.cell(r, 1).value).strip().lower() == "key":
            return r
    return 1


def read_template(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    hr = find_header(ws)
    # map year columns
    year_cols = {}
    for c in range(4, ws.max_column + 1):
        h = str(ws.cell(hr, c).value or "").strip().upper().replace("FY", "")
        if h.isdigit():
            year_cols[int(h)] = c

    metrics = {}
    all_keys = set()
    for r in range(hr + 1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if not key:
            continue
        key = str(key).strip()
        all_keys.add(key)
        label = ws.cell(r, 2).value
        unit = (ws.cell(r, 3).value or "").strip() if ws.cell(r, 3).value else ""
        by_year = {}
        for yr, c in year_cols.items():
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            try:
                num = float(v)
            except (TypeError, ValueError):
                continue
            by_year[str(yr)] = int(num) if num == int(num) else num
        latest = None
        if by_year:
            ly = max(int(y) for y in by_year)
            latest = {"year": ly, "value": by_year[str(ly)]}
        metrics[key] = {"label": label, "unit": unit, "by_year": by_year, "latest": latest}

    fam = "licensing" if all_keys & LICENSING_KEYS else ("industry" if all_keys & INDUSTRY_KEYS else "unknown")
    meta = FAMILY_META.get(fam, {"family": fam, "track": "Track 1 - internal", "source": "GT internal intake"})
    return {"metric_provenance": meta, "metrics": metrics}


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    outdir = Path(__file__).resolve().parent.parent / "output"
    outdir.mkdir(parents=True, exist_ok=True)
    for arg in sys.argv[1:]:
        p = Path(arg)
        if not p.exists():
            print(f"  skip (not found): {p}", file=sys.stderr); continue
        summary = read_template(p)
        fam = summary["metric_provenance"]["family"]
        out = outdir / f"internal_{fam}.json"
        out.write_text(json.dumps(summary, indent=2))
        n_filled = sum(1 for m in summary["metrics"].values() if m["latest"])
        print(f"{p.name} -> {out.name}  ({fam}; {n_filled}/{len(summary['metrics'])} metrics have data)")


if __name__ == "__main__":
    main()
